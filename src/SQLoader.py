import sqlite3
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from table_instructions import table_instructions


def extract_from_xlsx(wb, sheet_name, rectangle, column_names, transpose=False):
    '''
    Extracts data from an Excel worksheet and returns it as a pandas DataFrame.
    Parameters:
        wb: openpyxl Workbook object
        sheet_name: Name of the worksheet to extract data from
        rectangle: Cell range in A1 notation (e.g., "A1:C10")
        column_names: List of chosen column names for the DataFrame -> same as in sql table  (mostly the same but slugged, few exceptions)
        transpose: Boolean indicating whether to transpose the table
    '''
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(rectangle)

    rows = ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )

    data = list(rows)

    if transpose:
        data = np.array(data).T.tolist()

    return pd.DataFrame(data, columns=column_names)


def load_table_to_db(wb, table_name, config, conn):
    '''
    Loads a table from the Excel workbook into the SQLite database based on the provided configuration.
    '''
    df = extract_from_xlsx(
        wb,
        config["sheet_name"],
        config["rectangle"],
        config["column_names"],
        config["transpose"]
    )

    process = config.get("process")
    if process:
        df = process(df)

    # create table schema first
    conn.execute(f"DROP TABLE IF EXISTS {table_name}")
    conn.execute(config["schema"])

    # load data
    df.to_sql(table_name, conn, if_exists='append', index=False)


def transform_to_time_series(df, id_var, time_var, value_name):
    return df.melt(id_vars=[id_var], var_name=time_var, value_name=value_name).sort_values(by=[id_var, time_var]).reset_index(drop=True)


def build_player_devices_table(conn):
    conn.executescript("""
    CREATE TABLE player_devices AS
    SELECT 
        player_id,
        ev1_model,
        ev2_model,
        has_pv,
        has_ess
    FROM player_pv_ess
    LEFT JOIN player_evs USING(player_id);

    DROP TABLE player_pv_ess;
    DROP TABLE player_evs;
    """)

# make these importable in other files:
root_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
excel_file_path = os.path.join(root_dir, "data", "A.xlsx")
db_path = os.path.join(root_dir, "db", "energy.db")




def load_data_to_db():
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)

    wb = load_workbook(excel_file_path, data_only=True)

    # Load tables
    print("Loading tables...")
    for name in table_instructions.keys():  #tables_to_load:
        cfg = table_instructions[name]
        # print(f"Loading {name} ...")
        try:
            load_table_to_db(wb, name, cfg, conn)
        except Exception as e:
            print(f"Error loading {name}: {e}")

    print("Tables loaded.")

    # combine pv_ess table and evs table into player_devices
    build_player_devices_table(conn)

    conn.commit()
    conn.close()


if __name__ == "__main__":
    load_data_to_db()