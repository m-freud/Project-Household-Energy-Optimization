import sqlite3
import pandas as pd
import numpy as np
import sys
from pathlib import Path
import datetime as dt
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from influxdb_client.client.influxdb_client import InfluxDBClient
from influxdb_client.client.write.point import Point
from influxdb_client.client.write_api import WriteOptions
from influxdb_client.client.write_api import SYNCHRONOUS

from table_config import table_instructions

load_dotenv()

# file paths
ROOT_DIR = Path(__file__).parent.parent.parent
EXCEL_FILE_PATH = ROOT_DIR / "data" / "A.xlsx"
DB_PATH = ROOT_DIR / "db" / "energy.db"

# config file import
sys.path.append(str(ROOT_DIR))
from config import Config

# sqlite connection
conn = sqlite3.connect(DB_PATH)

# Influx connection
INFLUX_TOKEN = Config.INFLUX_TOKEN
INFLUX_URL = Config.INFLUX_URL
INFLUX_ORG = Config.INFLUX_ORG
INFLUX_BUCKET = Config.INFLUX_BUCKET
assert INFLUX_URL and INFLUX_TOKEN and INFLUX_ORG and INFLUX_BUCKET

client = InfluxDBClient(url=INFLUX_URL, token=INFLUX_TOKEN, org=INFLUX_ORG)
write_api = client.write_api(write_options=SYNCHRONOUS)
buckets_api = client.buckets_api()

if not buckets_api.find_bucket_by_name(INFLUX_BUCKET):
    buckets_api.create_bucket(bucket_name=INFLUX_BUCKET, org=INFLUX_ORG)


def extract_df_from_xlsx(wb, sheet_name, rectangle, column_names, transpose=False):
    '''
    Extracts data from an Excel worksheet and returns it as a pandas DataFrame.
    Parameters:
        wb: openpyxl Workbook object
        sheet_name: Name of the worksheet to extract data from
        rectangle: Cell range in A1 notation (e.g., "A1:C10")
        column_names: List of chosen column names for the DataFrame -> same as in sql table  (mostly the same but slugged, few exceptions)
        transpose: Boolean indicating whether to transpose the table
    '''
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(rectangle)

    rows = ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )

    data = list(rows)

    if transpose:
        data = np.array(data).T.tolist()

    return pd.DataFrame(data, columns=column_names)


def period_to_epoch(period):
    seconds = period * 15 * 60
    base = dt.datetime(2000,1,1,0,0)
    ts = base + dt.timedelta(seconds=seconds)
    return ts


def convert_to_timeseries(df):
    '''
    Converts a DataFrame from wide format to long format suitable for Influx.
    Assumes wide format (periods x players).
    '''
    # time conversion
    df["timestamp"] = df["period"].apply(period_to_epoch)

    # melt to long format
    df = df.melt(
        id_vars=["timestamp"],
        var_name="player_id",
        value_name="value").sort_values(by=["player_id", "timestamp"]).reset_index(drop=True)
    
    return df


def load_to_influx(df, table_name):
    '''
    Loads a DataFrame into InfluxDB as a time series.
    Assumes wide format (periods x players).
    '''
    assert INFLUX_BUCKET

    df = convert_to_timeseries(df)

    write_api.write(
        bucket=INFLUX_BUCKET,
        record=df,
        data_frame_measurement_name=table_name,
        data_frame_field_name="value",
        data_frame_tag_columns=["player_id"],
        data_frame_timestamp_column="timestamp")


def load_to_sqlite(df, table_name, config):
    process = config.get("process")
    if process:
        df = process(df)

    # create table schema first
    conn.execute(f"DROP TABLE IF EXISTS {table_name}")
    conn.execute(config["schema"])

    # load data
    df.to_sql(table_name, conn, if_exists='append', index=False)


def load_table_to_db(wb, table_name, config):
    '''
    Loads a table from the Excel workbook into SQLite or InfluxDB based on the provided configuration.
    '''
    df = extract_df_from_xlsx(
        wb,
        config["sheet_name"],
        config["rectangle"],
        config["df_column_names"],
        config["transpose"]
    )

    if config.get("time_series"):
        load_to_influx(df, table_name)
    else:
        load_to_sqlite(df, table_name, config)


def load_all_tables(wb, table_instructions):
    for table_name, config in table_instructions.items():
        print(f"Loading table: {table_name}...")
        load_table_to_db(wb, table_name, config)
        print("done!")


# load workbook
wb = load_workbook(EXCEL_FILE_PATH)

if __name__ == "__main__":
    wb = load_workbook(EXCEL_FILE_PATH)

    for table_name, config in table_instructions.items():
        print(f"Loading table: {table_name}...")
        load_table_to_db(wb, table_name, config)

    print("done!")

    conn.close()