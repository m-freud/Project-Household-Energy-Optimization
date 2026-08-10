import pandas as pd
import numpy as np
# paste this to enable src. imports
from pathlib import Path
import sys

# find the repository root that contains 'src'
repo_root = next((p for p in Path.cwd().resolve().parents if (p / "src").exists()), "")
sys.path.insert(0, str(repo_root))
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from src.ingestion.table_config import table_config
from src.sqlite_connection import sqlite_conn
from src.config import Config


def extract_df_from_xlsx(wb, sheet_name, rectangle, column_names, transpose=False):
    '''
    Extracts data from an Excel worksheet and returns it as a pandas DataFrame.
    Parameters:
        wb: openpyxl Workbook object
        sheet_name: Name of the worksheet to extract data from
        rectangle: Cell range in A1 notation (e.g., "A1:C10")
        column_names: List of chosen column names for the DataFrame -> same as in sql table  (mostly the same but slugged, few exceptions)
        transpose: Boolean indicating whether to transpose the table
    '''
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(rectangle)

    rows = ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )

    data = list(rows)

    if transpose:
        data = np.array(data).T.tolist()

    return pd.DataFrame(data, columns=column_names)


def load_to_sqlite(df, table_name, config):
    # apply optional processing function
    process = config.get("process")
    if process:
        df = process(df)

    # drop table if exists
    sqlite_conn.execute(f"DROP TABLE IF EXISTS {table_name}")

    # load data
    df.to_sql(table_name, sqlite_conn, if_exists='append', index=False)


def load_table_to_db(wb, table_name, config):
    '''
    Loads a table from the Excel workbook into SQLite or InfluxDB based on the provided configuration.
    '''
    df = extract_df_from_xlsx(
        wb,
        config["sheet_name"],
        config["rectangle"],
        config["df_column_names"],
        config.get("transpose", False)
    )

    load_to_sqlite(df, table_name, config)


def _extract_home_charge_from_limits(wb) -> pd.DataFrame:
    limits_config = table_config["fixed_costs"]
    limits_df = extract_df_from_xlsx(
        wb,
        limits_config["sheet_name"],
        limits_config["rectangle"],
        limits_config["df_column_names"],
        limits_config.get("transpose", False),
    )
    home_charge_df = limits_df[["player_id", "premium_charger_edp_capacity"]].copy()
    home_charge_df = home_charge_df.rename(
        columns={"premium_charger_edp_capacity": "charge_home"}
    )
    return home_charge_df


def enrich_ev_charge_rates(wb):
    home_charge_df = _extract_home_charge_from_limits(wb)
    home_charge_rows = []
    for player_id, charge_home in home_charge_df.itertuples(index=False):
        if pd.isna(player_id):
            continue
        parsed_charge_home = None if pd.isna(charge_home) else float(charge_home)
        home_charge_rows.append((int(player_id), parsed_charge_home))

    sqlite_conn.execute("DROP TABLE IF EXISTS temp_home_charge")
    sqlite_conn.execute(
        "CREATE TEMP TABLE temp_home_charge (player_id INTEGER PRIMARY KEY, charge_home REAL)"
    )
    sqlite_conn.executemany(
        "INSERT OR REPLACE INTO temp_home_charge (player_id, charge_home) VALUES (?, ?)",
        home_charge_rows,
    )

    # Add max charge rates for ev tables, for home and station
    for ev_table in ("ev1", "ev2"):
        existing_columns = {
            row[1] for row in sqlite_conn.execute(f"PRAGMA table_info({ev_table})").fetchall()
        }
        if "charge_home" not in existing_columns:
            sqlite_conn.execute(f"ALTER TABLE {ev_table} ADD COLUMN charge_home REAL")
        if "charge_station" not in existing_columns:
            sqlite_conn.execute(f"ALTER TABLE {ev_table} ADD COLUMN charge_station REAL")
        if "charge_slowest" not in existing_columns:
            sqlite_conn.execute(f"ALTER TABLE {ev_table} ADD COLUMN charge_slowest REAL")

        sqlite_conn.execute(
            f'''
            UPDATE {ev_table}
            SET
                charge_home = (
                    SELECT temp.charge_home
                    FROM temp_home_charge AS temp
                    WHERE temp.player_id = {ev_table}.player_id
                ),
                charge_station = COALESCE(charge, station_max_charge)
            '''
        )

        sqlite_conn.execute(
            f'''
            UPDATE {ev_table}
            SET
                charge_slowest = CASE
                    WHEN charge_home IS NULL THEN charge_station
                    WHEN charge_station IS NULL THEN charge_home
                    ELSE MIN(charge_home, charge_station)
                END
            '''
        )

    sqlite_conn.execute("DROP TABLE IF EXISTS temp_home_charge")
    sqlite_conn.commit()


def add_combined_ev_states():
    # ev status = 1 - ev_at_home + ev_at_charging_station, evaluated per household column
    for ev_num in ("1", "2"):
        source_home = f"ev{ev_num}_at_home"
        source_station = f"ev{ev_num}_at_charging_station"
        target_table = f"ev{ev_num}_status"

        home_df = pd.read_sql_query(f"SELECT * FROM {source_home}", sqlite_conn)
        station_df = pd.read_sql_query(f"SELECT * FROM {source_station}", sqlite_conn)

        if home_df.empty or station_df.empty:
            sqlite_conn.execute(f"DROP TABLE IF EXISTS {target_table}")
            pd.DataFrame().to_sql(target_table, sqlite_conn, if_exists="replace", index=False)
            continue

        if "period" not in home_df.columns or "period" not in station_df.columns:
            raise ValueError(f"{source_home} and {source_station} must contain a period column")

        status_df = home_df.copy()
        value_columns = [column for column in home_df.columns if column != "period"]

        for column in value_columns:
            status_df[column] = 1 - pd.to_numeric(home_df[column], errors="coerce") + pd.to_numeric(
                station_df[column], errors="coerce"
            )

        sqlite_conn.execute(f"DROP TABLE IF EXISTS {target_table}")
        status_df.to_sql(target_table, sqlite_conn, if_exists="replace", index=False)
        sqlite_conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{target_table}_period ON {target_table} (period)"
        )

    sqlite_conn.commit()


def load_all_tables(wb, table_instructions):
    for table_name, config in table_instructions.items():
        print(f"Loading table: {table_name} to sqlite...")
        load_table_to_db(wb, table_name, config)

    print("Enriching EV charge-rate metadata...")
    enrich_ev_charge_rates(wb)

    print("Building combined EV state tables...")
    add_combined_ev_states()
    
    print("tables loaded successfully!")


if __name__ == "__main__":
    wb = load_workbook(Config.EXCEL_FILE_PATH, data_only=True)

    load_all_tables(wb, table_config)

    sqlite_conn.close()
