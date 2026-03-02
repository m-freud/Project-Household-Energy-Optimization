import pandas as pd
import numpy as np
import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# paste this to enable src. imports
from pathlib import Path
import sys

# find the repository root that contains 'src'
repo_root = next((p for p in Path.cwd().resolve().parents if (p / "src").exists()), "")
sys.path.insert(0, str(repo_root))


from ingestion.table_load_config import table_ingestion_config
import src.connections as connections
from src.config import Config


# sqlite connection
sqlite_conn = sql_connection.create_sqlite_connection()


def extract_df_from_xlsx(wb, sheet_name, rectangle, column_names, transpose=False):
    '''
    Extracts data from an Excel worksheet and returns it as a pandas DataFrame.
    Parameters:
        wb: openpyxl Workbook object
        sheet_name: Name of the worksheet to extract data from
        rectangle: Cell range in A1 notation (e.g., "A1:C10")
        column_names: List of chosen column names for the DataFrame -> same as in sql table  (mostly the same but slugged, few exceptions)
        transpose: Boolean indicating whether to transpose the table
    '''
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(rectangle)

    rows = ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )

    data = list(rows)

    if transpose:
        data = np.array(data).T.tolist()

    return pd.DataFrame(data, columns=column_names)


def load_to_sqlite(df, table_name, config):
    # do something with the table before loading if needed
    process = config.get("process")
    if process:
        df = process(df)

    # create table schema first
    sqlite_conn.execute(f"DROP TABLE IF EXISTS {table_name}")

    if len(config.get("schema", "").strip()) > 0:
        sqlite_conn.execute(config["schema"])

    # load data
    df.to_sql(table_name, sqlite_conn, if_exists='append', index=False)


def load_table_to_db(wb, table_name, config):
    '''
    Loads a table from the Excel workbook into SQLite or InfluxDB based on the provided configuration.
    '''
    df = extract_df_from_xlsx(
        wb,
        config["sheet_name"],
        config["rectangle"],
        config["df_column_names"],
        config["transpose"]
    )

    load_to_sqlite(df, table_name, config)


def compute_and_store_averages():
    """
    Compute average profiles from SQLite and store them back to SQLite.
    """
    for measurement in ["pv_gen", "base_load"]:
        print(f"Computing average for {measurement}...")

        try:
            avg_df = pd.read_sql_query(
                f'''
                SELECT period, AVG(value) AS value
                FROM {measurement}
                GROUP BY period
                ORDER BY period
                ''',
                sqlite_conn
            )
        except Exception as exc:
            print(f"  Could not compute average for {measurement}: {exc}")
            continue

        if avg_df.empty:
            print(f"  No data found for {measurement}, skipping...")
            continue

        avg_table_name = f"{measurement}_avg"
        avg_df.to_sql(avg_table_name, sqlite_conn, if_exists='replace', index=False)

        max_val = avg_df['value'].max()
        if pd.notna(max_val) and max_val > 0:
            norm_df = avg_df.copy()
            norm_df['value'] = norm_df['value'] / max_val
            norm_table_name = f"{measurement}_avg_norm"
            norm_df.to_sql(norm_table_name, sqlite_conn, if_exists='replace', index=False)

        sqlite_conn.commit()
        print(f"  Stored {avg_table_name} ({len(avg_df)} timesteps)")


def load_all_tables(wb, table_instructions):
    for table_name, config in table_instructions.items():
        destination = "InfluxDB" if config.get("time_series") else "SQLite"
        print(f"Loading table: {table_name} to {destination}...")
        load_table_to_db(wb, table_name, config)
    
    print("\nComputing and storing average profiles...")
    compute_and_store_averages()
    
    print("done!")


if __name__ == "__main__":
    wb = load_workbook(Config.EXCEL_FILE_PATH, data_only=True)

    load_all_tables(wb, table_ingestion_config)

    sqlite_conn.close()
