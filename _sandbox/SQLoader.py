import sqlite3
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def extract_from_xlsx(wb, sheet_name, rectangle, column_names, transpose=False):
    '''
    Extracts data from an Excel worksheet and returns it as a pandas DataFrame.
    Parameters:
        wb: openpyxl Workbook object
        sheet_name: Name of the worksheet to extract data from
        rectangle: Cell range in A1 notation (e.g., "A1:C10")
        column_names: List of chosen column names for the DataFrame -> same as in sql table  (mostly the same but slugged, few exceptions)
        transpose: Boolean indicating whether to transpose the table
    '''
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(rectangle)

    rows = ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )

    data = list(rows)

    if transpose:
        data = np.array(data).T.tolist()

    return pd.DataFrame(data, columns=column_names)


def load_table_to_db(wb, table_name, config, conn):
    '''
    Loads a table from the Excel workbook into the SQLite database based on the provided configuration.
    '''
    df = extract_from_xlsx(
        wb,
        config["sheet_name"],
        config["rectangle"],
        config["column_names"],
        config["transpose"]
    )

    process = config.get("process")
    if process:
        df = process(df)

    # create table schema first
    conn.execute(f"DROP TABLE IF EXISTS {table_name}")
    conn.execute(config["schema"])

    # load data
    df.to_sql(table_name, conn, if_exists='append', index=False)
    print(f"Loaded table {table_name}.")


def transform_to_time_series(df, id_var, time_var, value_name):
    return df.melt(id_vars=[id_var], var_name=time_var, value_name=value_name).sort_values(by=[id_var, time_var]).reset_index(drop=True)


def build_player_devices_table(conn):
    conn.executescript("""
    CREATE TABLE player_devices AS
    SELECT 
        player_id,
        ev1_model,
        ev2_model,
        has_pv,
        has_ess
    FROM player_pv_ess
    LEFT JOIN player_evs USING(player_id);

    DROP TABLE player_pv_ess;
    DROP TABLE player_evs;
    """)


period_cols = ",\n".join([f"            period_{i} INTEGER" for i in range(1, 97)])

table_configs = {
    "player_pv_ess": {
        "sheet_name": "General Information",
        "rectangle": "A5:C254",
        "column_names": ["player_id", "has_pv", "has_ess"],
        "schema": """
        CREATE TABLE IF NOT EXISTS player_pv_ess (
            player_id INTEGER PRIMARY KEY,
            has_pv INTEGER,
            has_ess INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "total": {
        "sheet_name": "General Information",
        "rectangle": "E4:F6",
        "column_names": ["unit", "count"],
        "schema": """
        CREATE TABLE IF NOT EXISTS total (
            unit TEXT PRIMARY KEY,
            count INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "contractual_power_terms": {
        "sheet_name": "General Information",
        "rectangle": "E14:G24",
        "column_names": ["cp_kva", "price_per_day", "count"],
        "schema": """
        CREATE TABLE IF NOT EXISTS contractual_power_terms (
            cp_kva REAL PRIMARY KEY,
            price_per_day REAL,
            count INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "time_of_use_tariff_3": {
        "sheet_name": "General Information",
        "rectangle": "F28:G31",
        "column_names": ["cp_kva", "on_peak_eur", "mid_peak_eur", "off_peak_eur"],
        "schema": """
        CREATE TABLE IF NOT EXISTS time_of_use_tariff_3 (
            cp_kva TEXT PRIMARY KEY,
            on_peak_eur REAL,
            mid_peak_eur REAL,
            off_peak_eur REAL
        )""",
        "transpose": True,
        "process": lambda df: df.assign(
            cp_kva=df['cp_kva'].replace({
                "<=20,7": "leq_20_7",
                ">20,7": "gt_20_7"
            })
        )
    },
    # "time_of_use_tariff_4" : {    ????
    #     "sheet_name": "General Information",
    #     "rectangle": "E35:H37",
    #     "column_names": ["super_off_peak", "off_peak", "mid_peak", "on_peak"],
    #     "schema": ?
    #     "transpose": False,
    #     "process": """transform into table period vs peak_type. Where are the prices?"""
    # },
    "energy_storage_system_models": {
        "sheet_name": "General Information",
        "rectangle": "K5:Q12",
        "column_names": ["type", "capacity_kW", "charge_kW", "Discharge_kW", "efficiency_percent", "model", "units"],
        "schema": """
        CREATE TABLE IF NOT EXISTS energy_storage_system_models (
            type INTEGER PRIMARY KEY,
            capacity_kW REAL,
            charge_kW REAL,
            discharge_kW REAL,
            efficiency_percent REAL,
            model TEXT,
            units INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "premium_charger_edp": {
        "sheet_name": "General Information",
        "rectangle": "K17:M18",
        "column_names": ["type", "capacity_kW", "units"],
        "schema": """
        CREATE TABLE IF NOT EXISTS premium_charger_edp (
                    type INTEGER PRIMARY KEY,
                    capacity_kW REAL,
                    units INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "regular_ev_models": {
        "sheet_name": "General Information",
        "rectangle": "T5:AB19",
        "column_names" : ["model_id", "brand", "model_name", "capacity_kW", "charge_kW", "discharge_kW", "efficiency_percent", "consumption_Wh_km", "units"],
        "schema": """
        CREATE TABLE IF NOT EXISTS regular_ev_models (
            model_id INTEGER PRIMARY KEY,
            brand TEXT,
            model_name TEXT,
            capacity_kW REAL,
            charge_kW REAL,
            discharge_kW REAL,
            efficiency_percent REAL,
            consumption_Wh_km REAL,
            units INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "premium_ev_models": {
        "sheet_name": "General Information",
        "rectangle": "T24:AB33",
        "column_names" : ["model_id", "brand", "model_name", "capacity_kW", "charge_kW", "discharge_kW", "efficiency_percent", "consumption_Wh_km", "units"],
        "schema": """
        CREATE TABLE IF NOT EXISTS premium_ev_models (
            model_id INTEGER PRIMARY KEY,
            brand TEXT,
            model_name TEXT,
            capacity_kW REAL,
            charge_kW REAL,
            discharge_kW REAL,
            efficiency_percent REAL,
            consumption_Wh_km REAL,
            units INTEGER
        )""",
        "transpose": False,
        "process": None
    },
    "ev_departures_arrivals": {
        "sheet_name": "General Information",
        "rectangle": "AE6:AK33",
        "column_names": ["period", "departures_ev1", "arrivals_ev1", "departures_ev2", "arrivals_ev2", "departures_total", "arrivals_total"],
        "schema": """
        CREATE TABLE IF NOT EXISTS ev_departures_arrivals (
            period INTEGER,
            departures_ev1 INTEGER,
            arrivals_ev1 INTEGER,
            departures_ev2 INTEGER,
            arrivals_ev2 INTEGER,
            departures_total INTEGER,
            arrivals_total INTEGER
        )""",
        "transpose": False,
        "process": lambda df: df.dropna(how="all")  # drop empty rows
    },
    "mobile_ev_chargers": {
        "sheet_name": "MOBIe EV chargers",
        "rectangle": "A3:O86",
        "column_names": ["charger_id", "charger_parent_uid", "charger_uid", "type_of_charging", "state_of_charger","city",	"address", "operator", "capacity_kW", "voltage_level", "price_eur_per_charge",	"price_eur_per_minute",	"price_eur_per_kWh","price_eur_per_h",	"price_eur_per_kWh_2"],
        "schema": """
        CREATE TABLE IF NOT EXISTS mobile_ev_chargers (
            charger_id INTEGER PRIMARY KEY,
            charger_parent_uid INTEGER,
            charger_uid TEXT,
            type_of_charging TEXT,
            state_of_charger TEXT,
            city TEXT,
            address TEXT,
            operator TEXT,
            capacity_kW REAL,
            voltage_level TEXT,
            price_eur_per_charge REAL,
            price_eur_per_minute REAL,
            price_eur_per_kWh REAL,
            price_eur_per_h REAL,
            price_eur_per_kWh_2 REAL
        )""",
        "transpose": False,
        "process": None
    },
    "player_evs": {
        "sheet_name": "EVs",
        "rectangle": "B1:IQ3",
        "column_names": ["player_id", "ev1_model", "ev2_model"],
        "schema": """
        CREATE TABLE IF NOT EXISTS player_evs (
            player_id INTEGER PRIMARY KEY,
            ev1_model TEXT,
            ev2_model TEXT
        )""",
        "transpose": True,
        "process": None      # we combine this with player_devices later
    },
    "ev1_data": {
        "sheet_name": "EVs",
        "rectangle": "B6:IQ25",
        "column_names": [
            "player_id",
            "model_id",
            "capacity_kw",
            "charge_kw",
            "discharge_kw",
            "efficiency",
            "initial_battery_level_kw",
            "final_battery_level_kw",
            "consumption_wh_per_km",
            "departure_period",
            "arrival_period",
            "distance_km",
            "consumption_kwh",
            "morning_trip_duration_periods",
            "afternoon_trip_duration_periods",
            "trip_duration_periods",
            "charger_type",
            "public_charger",
            "power_kw",
            "price_at_public_charge_station_eur"],
        "schema": """
        CREATE TABLE IF NOT EXISTS ev1_data (
            player_id INTEGER PRIMARY KEY,
            model_id INTEGER,
            capacity_kw REAL,
            charge_kw REAL,
            discharge_kw REAL,
            efficiency REAL,
            initial_battery_level_kw REAL,
            final_battery_level_kw REAL,
            consumption_wh_per_km REAL,
            departure_period INTEGER,
            arrival_period INTEGER,
            distance_km REAL,
            consumption_kwh REAL,
            morning_trip_duration_periods INTEGER,
            afternoon_trip_duration_periods INTEGER,
            trip_duration_periods INTEGER,
            charger_type INTEGER,
            public_charger INTEGER,
            power_kw REAL,
            price_at_public_charge_station_eur REAL
        )""",
        "transpose": True,
        "process": None
    },
    "ev2_data": {
        "sheet_name": "EVs",
        "rectangle": "B27:IQ46",
        "column_names": [
            "player_id",
            "model_id",
            "capacity_kw",
            "charge_kw",
            "discharge_kw",
            "efficiency",
            "initial_battery_level_kw",
            "final_battery_level_kw",
            "consumption_wh_per_km",
            "departure_period",
            "arrival_period",
            "distance_km",
            "consumption_kwh",
            "morning_trip_duration_periods",
            "afternoon_trip_duration_periods",
            "trip_duration_periods",
            "charger_type",
            "public_charger",
            "power_kw",
            "price_at_public_charge_station_eur"],
        "schema": """
            CREATE TABLE IF NOT EXISTS ev2_data (
                player_id INTEGER PRIMARY KEY,
                model_id INTEGER,
                capacity_kw REAL,
                charge_kw REAL,
                discharge_kw REAL,
                efficiency REAL,
                initial_battery_level_kw REAL,
                final_battery_level_kw REAL,
                consumption_wh_per_km REAL,
                departure_period INTEGER,
                arrival_period INTEGER,
                distance_km REAL,
                consumption_kwh REAL,
                morning_trip_duration_periods INTEGER,
                afternoon_trip_duration_periods INTEGER,
                trip_duration_periods INTEGER,
                charger_type INTEGER,
                public_charger INTEGER,
                power_kw REAL,
                price_at_public_charge_station_eur REAL
            )""",
        "transpose": True,
        "process": None
    },
    "load": {
        "sheet_name": "Load",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS load (
                    player_id INTEGER,
                    period INTEGER,
                    load_kW REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "load_kW")
    },
    "pv": {
        "sheet_name": "PV",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "transpose": True,
        "schema": f"""
        CREATE TABLE IF NOT EXISTS pv (
                    player_id INTEGER,
                    period INTEGER,
                    pv_kW REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "pv_kW")
    },
    "bess": {
        "sheet_name": "BESS",
        "rectangle": "B1:IQ8",
        "column_names": ["player_id", "model_id", "capacity_kW", "charge_kW", "discharge_kW", "efficiency", "initial_kW", "final_kW"],
        "schema": """
        CREATE TABLE IF NOT EXISTS bess (
            player_id INTEGER PRIMARY KEY,
            model_id INTEGER,
            capacity_kW REAL,
            charge_kW REAL,
            discharge_kW REAL,
            efficiency REAL,
            initial_kW REAL,
            final_kW REAL
        )""",
        "transpose": True,
        "process": None
    },
    "buy_price": {
        "sheet_name": "Buy Price",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS buy_price (
                        player_id INTEGER,
                        period INTEGER,
                        buy_price_eur REAL,
                        PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "buy_price_eur")
    },
    "sell_price": {
        "sheet_name": "Sell Price",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS sell_price (
                    player_id INTEGER,
                    period INTEGER,
                    sell_price_eur REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "sell_price_eur")
    },
    "limits": {
        "sheet_name": "Limits",
        "rectangle": "B1:IQ10",
        "column_names": ["player_id", "power_buy_kW", "power_sell_kW", "fixed_costs_eur","_","initial_cp_kW","premium_charger_edp_capacity_kW","sum_kW","new_cp_level_kW","fixed_costs_2_eur"],
        "schema": """
        CREATE TABLE IF NOT EXISTS limits (
            player_id INTEGER PRIMARY KEY,
            power_buy_kW REAL,
            power_sell_kW REAL,
            fixed_costs_eur REAL,
            initial_cp_kW REAL,
            premium_charger_edp_capacity_kW REAL,
            sum_kW REAL,
            new_cp_level_kW REAL,
            fixed_costs_2_eur REAL
        )""",
        "transpose": True,
        "process": lambda df: df.drop(columns=["_"])
    },
    "ev1_buy_price": {
        "sheet_name": "EV1 Buy Price",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev1_buy_price (
                    player_id INTEGER,
                    period INTEGER,
                    ev1_buy_price_eur REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev1_buy_price_eur")
    },
    "ev2_buy_price": {
        "sheet_name": "EV2 Buy Price",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev2_buy_price (
                    player_id INTEGER,
                    period INTEGER,
                    ev2_buy_price_eur REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev2_buy_price_eur")
    },
    "ev1_load": {
        "sheet_name": "EV1 Load",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev1_load (
                    player_id INTEGER,
                    period INTEGER,
                    ev1_load REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev1_load")
    },
    "ev2_load": {
        "sheet_name": "EV2 Load",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev2_load (
                    player_id INTEGER,
                    period INTEGER,
                    ev2_load REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev2_load")
    },
    "ev1_max_discharge": {
        "sheet_name": "Max EV1 Dis-Charge",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev1_max_discharge (
                    player_id INTEGER,
                    period INTEGER,
                    ev1_max_discharge REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev1_max_discharge")
    },
    "ev2_max_discharge": {
        "sheet_name": "Max EV2 Dis-Charge",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev2_max_discharge (
                    player_id INTEGER,
                    period INTEGER,
                    ev2_max_discharge REAL,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev2_max_discharge")
    },
    "ev1_at_home": {
        "sheet_name": "EV1 at Home (x)",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev1_at_home (
                    player_id INTEGER,
                    period INTEGER,
                    ev1_at_home BOOLEAN,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev1_at_home")
    },
    "ev2_at_home": {
        "sheet_name": "EV2 at Home (x)",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev2_at_home (
                    player_id INTEGER,
                    period INTEGER,
                    ev2_at_home BOOLEAN,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev2_at_home")
    },
    "ev1_at_charging_station": {
        "sheet_name": "EV1 at Charging Station (x)",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev1_at_charging_station (
                    player_id INTEGER,
                    period INTEGER,
                    ev1_at_charging_station BOOLEAN,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev1_at_charging_station")
    },
    "ev2_at_charging_station": {
        "sheet_name": "EV2 at Charging Station (x)",
        "rectangle": "B1:IQ97",
        "column_names": ["player_id"] + [i for i in range(1, 96+1)],
        "schema": f"""
        CREATE TABLE IF NOT EXISTS ev2_at_charging_station (
                    player_id INTEGER,
                    period INTEGER,
                    ev2_at_charging_station BOOLEAN,
                    PRIMARY KEY (player_id, period)
        )""",
        "transpose": True,
        "process": lambda df: transform_to_time_series(df, "player_id", "period", "ev2_at_charging_station")
    }
}

tables_to_load = [
    # "players",
    # "total",
    # "contractual_power_terms",
    # "load",
]

if os.path.exists("energy.db"):
    os.remove("energy.db")

conn = sqlite3.connect("energy.db")

data_path = os.path.join(os.getcwd(), "data", "A.xlsx")
wb = load_workbook(data_path, data_only=True)

# Load tables
for name in table_configs.keys():  #tables_to_load:
    cfg = table_configs[name]
    print(f"Loading {name} ...")
    try:
        load_table_to_db(wb, name, cfg, conn)
    except Exception as e:
        print(f"Error loading {name}: {e}")

# combine pv_ess table and evs table into player_devices
build_player_devices_table(conn)

conn.commit()
conn.close()
